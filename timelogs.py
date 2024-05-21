# :coding: utf-8

import os
from pprint import pprint
import shotgun_api3 as sa
import datetime as dt
import openpyxl
import requests
from openpyxl.drawing.image import Image
from collections import OrderedDict
import uuid



sg = sa.Shotgun(
                'https://west.shotgunstudio.com',
                api_key='6h&ziahfuGbqdubxrxeyopinl',
                script_name = 'eventTrigger',
            )

mana_list = sg.find( 
            'CustomNonProjectEntity04',
            [
                ['sg_use_filed','is','use' ]
            ],
            ['code']
)


## Human Resource 
## TODO List
#   - 썸네일 위치값 표시
#   - 시작 기준으로 정렬, 매월 기주으로 프로젝트 중복되도 표기
#   - 샷 카운트가 샷인지 테스크인지 확인
#   - 샷작업 이외 항목들, 미팅, 알엔디 따로 표기 되도록 etc 종류별로 시간 표기

def main( 
            user , 
            s_year, s_month, s_day, e_year = None, 
            e_month = None, e_day = None, 
            save_path = os.path.expanduser('~') 
        ):
    started = dt.datetime( s_year, s_month, s_day )    
    ended   = dt.datetime( e_year, e_month, e_day )    
    
    expand_end = False

    if ended > dt.datetime( 2024, 5, 1 ):
        expand_end = ended
        ended = dt.datetime( 2024, 4, 30 )


    filters = [
                  ['created_at', 'greater_than', started ],
                  ['user.HumanUser.firstname', 'is', user ],
              ]
    if ended:
        filters.append( 
                ['created_at', 'less_than' , ended ]
        )

    result = sg.find( 
                    'TimeLog', filters,
                    [ 
                        'created_at','user', 'entity', 'duration','project', 
                        'user.HumanUser.department.Department.name',
                        'department',
                    ]
    )
    
    for x in result:
        x['thumb'] = down_thumb( x )
    
    ######################
    ## Column value
    ## 투입월 , 부서, 이름, 프로젝트명, 총투입시간, 샷 count, 샷 소요시간, rnd count ,소요시간,  meeting count, 소요시간....Thumb, 비고(thumb 위치 )    

    fin_data = OrderedDict()
    for _data in result:
        month = '{}.{}'.format( _data['created_at'].year , _data['created_at'].month )
        if month not in fin_data.keys():
            fin_data[ month ] = {}
        show = _data['project']['name']
        if show not in fin_data[month].keys() :
            fin_data[month][show] = {}

        entity = _data['entity']['name']
        if entity not in fin_data[month][show]:
            fin_data[month][show][entity] = [ ]

        fin_data[month][show][entity].append(  _data  )
    
    #return fin_data 
    wb = openpyxl.Workbook()

    sheet = wb.active
    sheet.title = '24.5월이전'

    col_list = [ 
            '투입월','부서', '이름', '프로젝트명', 'Shot Count', 'Task시간', '총투입시간',  'Meeting Count','Meeting 시간','CGIsup Count',
            'CGIsup 시간', 'RND', 'RND 시간', 'Out Work', 'Out Work 시간', 'Teamlead', 'Teamlead 시간', 'No info',
            'No info 시간', 'Previz', 'Previz 시간', 'Concept', 'Concetp 시간', 'Thumb', 'Thumb 비고'
    ]
    
    for num, col in enumerate( col_list):
        sheet.cell( row = 1 , column = num + 1 ).value = col

    row = 2
    for irow, month in enumerate( fin_data):
        for show in fin_data[ month ]:
            #for entity in fin_data[month][show]:
            _data = fin_data[month][show]
            etc_tasks = get_classify_task( _data )
            all_etc_task_duration = sum( [ etc_tasks[x][1] for x in etc_tasks ] ) / 60.0

            sheet.cell( row  , column = 1 ).value = month
            sheet.cell( row  , column = 2  ).value = _data[ list(_data.keys())[0] ][0]['user.HumanUser.department.Department.name']
            sheet.cell( row  , column = 3  ).value = user
            sheet.cell( row  , column = 4  ).value = show
            sheet.cell( row  , column = 5  ).value = len( get_all_tasks(_data ) )
            sheet.cell( row  , column = 6  ).value = round( get_all_timelogs(_data ) / 60.0 - all_etc_task_duration , 1 )
            sheet.cell( row  , column = 7  ).value = get_all_timelogs(_data ) / 60.0
            if 'Meeting' in list( etc_tasks.keys() ):
                sheet.cell( row  , column = 8  ).value = etc_tasks['Meeting'][0]
                sheet.cell( row  , column = 9  ).value = etc_tasks['Meeting'][1] / 60.0
            if 'CGI_sup' in list( etc_tasks.keys() ):
                sheet.cell( row  , column = 10 ).value = etc_tasks['CGI_sup'][0]
                sheet.cell( row  , column = 11 ).value = etc_tasks['CGI_sup'][1] / 60.0
            if 'rnd' in list( etc_tasks.keys() ):
                sheet.cell( row  , column = 12 ).value = etc_tasks['rnd'][0]
                sheet.cell( row  , column = 13 ).value = etc_tasks['rnd'][1] / 60.0
            if 'OutWork' in list( etc_tasks.keys() ):
                sheet.cell( row  , column = 14 ).value = etc_tasks['OutWork'][0]
                sheet.cell( row  , column = 15 ).value = etc_tasks['OutWork'][1] / 60.0
            if 'TeamLead' in list( etc_tasks.keys() ):
                sheet.cell( row  , column = 16 ).value = etc_tasks['TeamLead'][0]
                sheet.cell( row  , column = 17 ).value = etc_tasks['TeamLead'][1] / 60.0
            if 'NoInfo' in list( etc_tasks.keys() ):
                sheet.cell( row  , column = 18 ).value = etc_tasks['NoInfo'][0]
                sheet.cell( row  , column = 19 ).value = etc_tasks['NoInfo'][1] / 60.0
            if 'Previz' in list( etc_tasks.keys() ):
                sheet.cell( row  , column = 20 ).value = etc_tasks['Previz'][0]
                sheet.cell( row  , column = 21 ).value = etc_tasks['Previz'][1] / 60.0
            if 'Concept' in list( etc_tasks.keys() ):
                sheet.cell( row  , column = 22 ).value = etc_tasks['Concept'][0]
                sheet.cell( row  , column = 23 ).value = etc_tasks['Concept'][1] / 60.0


            thumb_tl = sel_thumb( _data )
            thumb_path = thumb_tl['thumb']
            if thumb_path:
                img = Image( thumb_path )
                #width , height = img.width, img.height
                width, height = get_col_size( img.width, img.height )
                sheet.column_dimensions['X'].width = width
                sheet.row_dimensions[row].height = height
                sheet.add_image( img, 'X' + str( row ) )
                sheet.cell( row  , column = 25 ).value = thumb_tl['location']['entity_name']
            else:
                sheet.cell( row = row , column = 24 ).value = 'No Image'
            row += 1
                
    if expand_end:
        
        filters = [
                      ['created_at', 'greater_than', dt.datetime( 2024,5,1) ],
                      ['user.HumanUser.firstname', 'is', user ],
                      ['created_at', 'less_than' , expand_end ]
                  ]

        result = sg.find( 
                        'TimeLog', filters,
                        [ 
                            'created_at','user', 'entity', 'duration','project', 
                            'user.HumanUser.department.Department.name',
                            'department',
                        ]
        )

        sheet1 = wb.create_sheet( '24.5월 이후', 1 )
        for x in result:
            x['thumb'] = down_thumb( x )

        fin_data = OrderedDict()
        for _data in result:
            month = '{}.{}'.format( _data['created_at'].year , _data['created_at'].month )
            if month not in fin_data.keys():
                fin_data[ month ] = {}
            show = _data['project']['name']
            if show not in fin_data[month].keys() :
                fin_data[month][show] = {}

            entity = _data['entity']['name']
            if entity not in fin_data[month][show]:
                fin_data[month][show][entity] = [ ]

            fin_data[month][show][entity].append(  _data  )
        col_list = [ 
                '투입월','부서', '이름', '프로젝트명', 'Shot Count', 'Task시간', '총투입시간',  
                'Work','Work 시간','Management','Management 시간', 'Dayoff','Dayoff 시간', 'H_Dayoff', 'H_Dayoff 시간',
                'Thumb', 'Thumb 비고'
        ]
        
        for num, col in enumerate( col_list):
            sheet1.cell( row = 1 , column = num + 1 ).value = col

        row = 2
        for irow, month in enumerate( fin_data):
            for show in fin_data[ month ]:
                #for entity in fin_data[month][show]:
                _data = fin_data[month][show]
                etc_tasks = get_classify_task( _data , new_period = True )
                all_etc_task_duration = sum( [ etc_tasks[x][1] for x in etc_tasks ] ) / 60.0

                sheet1.cell( row  , column = 1  ).value = month
                sheet1.cell( row  , column = 2  ).value = _data[ list(_data.keys())[0] ][0]['user.HumanUser.department.Department.name']
                sheet1.cell( row  , column = 3  ).value = user
                sheet1.cell( row  , column = 4  ).value = show
                sheet1.cell( row  , column = 5  ).value = len( get_all_tasks(_data ) )
                sheet1.cell( row  , column = 6  ).value = get_all_timelogs(_data ) / 60.0 - all_etc_task_duration
                sheet1.cell( row  , column = 7  ).value = get_all_timelogs(_data ) / 60.0
                if 'Work' in list( etc_tasks.keys() ):
                    sheet1.cell( row  , column = 8  ).value = etc_tasks['Work'][0]
                    sheet1.cell( row  , column = 9  ).value = etc_tasks['Work'][1] / 60.0
                if 'Management' in list( etc_tasks.keys() ):
                    sheet1.cell( row  , column = 10  ).value = etc_tasks['Management'][0]
                    sheet1.cell( row  , column = 11  ).value = etc_tasks['Management'][1] / 60.0
                if 'Dayoff' in list( etc_tasks.keys() ):
                    sheet1.cell( row  , column = 12  ).value = etc_tasks['Dayoff'][0]
                    sheet1.cell( row  , column = 13  ).value = etc_tasks['Dayoff'][1] / 60.0
                if 'H_Dayoff' in list( etc_tasks.keys() ):
                    sheet1.cell( row  , column = 14  ).value = etc_tasks['H_Dayoff'][0]
                    sheet1.cell( row  , column = 15  ).value = etc_tasks['H_Dayoff'][1] / 60.0

                thumb_tl = sel_thumb( _data )
                thumb_path = thumb_tl['thumb']
                if thumb_path:
                    img = Image( thumb_path )
                    #width , height = img.width, img.height
                    width, height = get_col_size( img.width, img.height )
                    sheet1.column_dimensions['P'].width = width
                    sheet1.row_dimensions[row].height = height
                    sheet1.add_image( img, 'P' + str( row ) )
                    sheet1.cell( row  , column = 17 ).value = thumb_tl['location']['entity_name']
                else:
                    sheet1.cell( row = row , column = 16 ).value = 'No Image'
                row += 1

        
        
    
    #xlsx_file = save_path #+ '/tmp/hr_{}.xlsx'.format( user )

    
    #xlsx_file = '/home/w10137/tmp/hr_{}.xlsx'.format( user )
    xlsx_file = save_path + os.sep + user + dt.datetime.now().strftime( '%y%m%d' )
    wb.save( xlsx_file )

    if os.path.exists( xlsx_file ):
        print( xlsx_file )
    else:
        print( 'Error' )

    return fin_data


def get_all_tasks( _data ):
    result = []
    for entity in _data:
        result.extend( _data[ entity ] )
    return result        

def get_all_timelogs( _data ):
    tasks = get_all_tasks( _data )
    return sum( [ x['duration'] for  x in tasks ] )

def sel_thumb( _data ):
    timelogs = get_all_tasks( _data )
    temp = [ x for x in timelogs if x['entity']['name'] not in mana_list ]
    ordered = sorted( temp, key = lambda x : x['duration'], reverse = True )
    

    for x in ordered:
        if x['thumb']:
        

            key = [ 'project.Project.name' ]
            if x['entity']['type'] == 'Asset':
                key.append( 'code' )
            elif x['entity']['type'] == 'Shot':
                key.append( 'code' )
            elif x['entity']['type'] == 'Task':
                key.append( 'entity' )
                #key.append( 'entity.{}.code'.format(x['entity']]['type'] ) )

            location = sg.find_one( 
                x['entity']['type'],
                [
                    ['id', 'is' , x['entity']['id'] ],
                ],
                key
            )
            x['location'] = location

            if x['entity']['type'] != 'Task':
                task_entity = sg.find_one( 
                    [
                        ['id', 'is', location['entity']['id']]
                    ],
                    [ 'entity.{}.code'.format(x['entity']['type'] ) ] 
                )
                x['location']['name'] = task_entity[ 'entity.{}.code'.format(x['entity']['type'] ) ] 

            if location[ 'entity' ]['type'] in ['Shot', 'Asset']:
                x['location']['entity_name'] = location['entity']['name']
                
#            if 'entity.{}.code'.format( x['entity']['entity']['type'] )  in key:
#                x['location']['code'] = location['entity.{}.code'.format( x['entity']['entity']['type'] ) ] 
            return x
    return { 'thumb' : '' }
    

def get_classify_task( _data, new_period = False ):
    exclude_mana = [ 'Work', 'Management', 'Dayoff', 'H_Dayoff' ]
    if new_period:
        mana_code_list = [ x['code']  for x in  mana_list if x['code'] in exclude_mana ]
    else:
        mana_code_list = [ x['code']  for x in  mana_list if x['code'] not in exclude_mana ]
    
    result = OrderedDict()
    for code in mana_code_list:
        result[code] = []


    for entity in _data:
        if entity in mana_code_list:
            result[ entity ].extend( _data[entity] )

    fin_data = {}
    for entity in result:
        fin_data[ entity ] = [ len( result[entity] ) , sum( [ x['duration'] for x in result[entity] ] ) ]                 
    
    return fin_data





def get_max( _iter ):
    _max = {}
    _max['duration'] = 0
    for  x in _iter:
        if _max['duration'] < x['duration']:
            _max = x
    return _max


def get_period( _iter ):
    start = {'created_at':0 }
    end = {'created_at':0 }
    for x in _iter:
        if start[ 'created_at'] == 0:
            start = x
        else: 
            if start['created_at'] > x['created_at' ]:
                start = x

        if end[ 'created_at' ] == 0:
            end = x
        else:
            if end['created_at'] < x['created_at' ]:
                end = x
    return start, end

def down_thumb( timelog ):
    entity = sg.find_one( 
                'Task',
                [ 
                    [ 'id' , 'is', timelog['entity']['id'] ]
                ], 
                [ 'image', 'content']
    )
    if not entity:
        return

    if not entity['image']:
        return ''        

    response = requests.get( entity['image'] )    

    image_save_path = os.path.expanduser( '~' ) + os.sep + 'tmp/thumb' + os.sep + str( timelog['entity']['id']  )
    image_save_path = image_save_path + '_' + str( uuid.uuid1() ).split('-')[0] + '.jpg'
    if not os.path.exists( os.path.dirname( image_save_path ) ):
        os.makedirs( os.path.dirname( image_save_path ) )

    response = requests.get( entity['image'] )

    if response.status_code == 200:
        with open( image_save_path , 'wb' ) as f:
            f.write( response.content )
    return image_save_path


def get_col_size( width, height ):
    col_width = width * 65.2 / 454.19
    row_height = height * 225.35 / 298.96
    return col_width , row_height


if __name__ == '__main__':
    start = dt.datetime.now()

    pprint( 
        main( '김상큼' , 2024, 1,1, 2024,4,30  )
    )

#    pprint( get_etc_task() )
    for _user in [ '김규남']:
        main( _user ) 
#    pprint( result )
#    print( '+' * 40 )
#    print( len( result )  )
#    print( '+' * 40 )
#    print( result.keys() )
#    print( '+' * 40 )
#    for x in result:
#        print( result[x].keys() )
#    print( '+' * 40 )
    end = dt.datetime.now()
    print( end - start )
